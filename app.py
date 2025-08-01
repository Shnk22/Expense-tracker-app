import streamlit as st
import pandas as pd
import os
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

FILE_PATH = "expenses.xlsx"

# ✅ Initialize Sheets
def initialize_sheets():
    if not os.path.exists(FILE_PATH):
        with pd.ExcelWriter(FILE_PATH, engine="openpyxl", mode="w") as writer:
            pd.DataFrame(columns=["Date", "Amount", "Category", "Notes", "Month"]).to_excel(writer, sheet_name="Expenses", index=False)
            pd.DataFrame(columns=["Date", "Medicine", "Quantity", "Cost", "Notes"]).to_excel(writer, sheet_name="Medicines", index=False)
            pd.DataFrame(columns=["Date", "Type", "Amount", "Frequency", "Notes"]).to_excel(writer, sheet_name="Investments", index=False)
        return

    wb = load_workbook(FILE_PATH)
    if "Medicines" not in wb.sheetnames:
        df = pd.DataFrame(columns=["Date", "Medicine", "Quantity", "Cost", "Notes"])
        with pd.ExcelWriter(FILE_PATH, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, sheet_name="Medicines", index=False)

    if "Investments" not in wb.sheetnames:
        df = pd.DataFrame(columns=["Date", "Type", "Amount", "Frequency", "Notes"])
        with pd.ExcelWriter(FILE_PATH, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, sheet_name="Investments", index=False)

initialize_sheets()

# ---------------- Sidebar ----------------
st.sidebar.title("📂 Menu")
menu = st.sidebar.radio("Go to:", ["💸 Expenses", "💊 Medicines", "💰 Investments"])

# ---------------- Expenses Functions ----------------
def load_expense_data():
    try:
        return pd.read_excel(FILE_PATH, sheet_name="Expenses")
    except:
        return pd.DataFrame(columns=["Date", "Amount", "Category", "Notes", "Month"])

def save_expense_data(new_row):
    df = load_expense_data()
    if not df.empty:
        df["Date"] = pd.to_datetime(df["Date"]).dt.date

    new_date = pd.to_datetime(new_row["Date"]).date()
    is_duplicate = (
        (df["Date"] == new_date) &
        (df["Amount"] == new_row["Amount"]) &
        (df["Category"] == new_row["Category"]) &
        (df["Notes"] == new_row["Notes"])
    ).any()
    if is_duplicate:
        st.warning("⚠️ This entry already exists.")
        return

    new_row["Date"] = new_date
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df["Month"] = pd.to_datetime(df["Date"]).dt.strftime('%B %Y')

    with pd.ExcelWriter(FILE_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name="Expenses", index=False)

    wb = load_workbook(FILE_PATH)
    ws = wb["Expenses"]
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    wb.save(FILE_PATH)

# ---------------- Medicines Functions ----------------
def load_medicines_data():
    try:
        return pd.read_excel(FILE_PATH, sheet_name="Medicines")
    except:
        return pd.DataFrame(columns=["Date", "Medicine", "Quantity", "Cost", "Notes"])

def save_medicine_data(new_row):
    df = load_medicines_data()
    new_row["Date"] = pd.to_datetime(new_row["Date"]).date()
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

    with pd.ExcelWriter(FILE_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name="Medicines", index=False)

# ---------------- Investments Functions ----------------
def load_investments_data():
    try:
        return pd.read_excel(FILE_PATH, sheet_name="Investments")
    except:
        return pd.DataFrame(columns=["Date", "Type", "Amount", "Frequency", "Notes"])

def save_investment_data(new_row):
    df = load_investments_data()
    new_row["Date"] = pd.to_datetime(new_row["Date"]).date()
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

    with pd.ExcelWriter(FILE_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name="Investments", index=False)

# ---------------- Expenses Tab ----------------
if menu == "💸 Expenses":
    st.title("💸 Expense Tracker")

    with st.form("expense_form"):
        exp_date = st.date_input("Date", value=date.today())
        amount = st.number_input("Amount (₹)", min_value=0.0, step=10.0)
        category = st.selectbox("Category", ["Food", "Transport", "Shopping", "Donation", "Bills", "Other"])
        notes = st.text_input("Notes (optional)")
        submit = st.form_submit_button("Add Expense")

    if submit and amount > 0:
        save_expense_data({"Date": exp_date, "Amount": amount, "Category": category, "Notes": notes})
        st.success("✅ Expense added successfully!")

    st.header("📊 Filtered Expenses by Month")
    df = load_expense_data()
    if not df.empty:
        df["Date"] = pd.to_datetime(df["Date"])
        df["Month"] = df["Date"].dt.strftime("%B %Y")
        months = df["Month"].unique()
        selected_month = st.selectbox("Select Month", sorted(months, reverse=True))
        filtered_df = df[df["Month"] == selected_month].copy().reset_index(drop=True)

        st.markdown("### Entries")
        for i, row in filtered_df.iterrows():
            cols = st.columns([2, 2, 2, 3, 1])
            cols[0].write(row["Date"].strftime("%Y-%m-%d"))
            cols[1].write(f"₹{row['Amount']}")
            cols[2].write(row["Category"])
            cols[3].write(row["Notes"])
            if cols[4].button("🗑", key=f"delete_{i}"):
                df = df.drop(index=i)
                with pd.ExcelWriter(FILE_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    df.to_excel(writer, sheet_name="Expenses", index=False)
                st.rerun()

        st.markdown(f"**Total Spent in {selected_month}: ₹{filtered_df['Amount'].sum():.2f}**")
    else:
        st.info("No expenses recorded yet.")

# ---------------- Medicines Tab ----------------
if menu == "💊 Medicines":
    st.title("💊 Medicines Tracker")

    with st.form("medicine_form"):
        med_date = st.date_input("Date of Purchase", value=date.today())
        med_name = st.text_input("Medicine Name")
        quantity = st.number_input("Quantity", min_value=1, step=1)
        cost = st.number_input("Cost (₹)", min_value=0.0, step=10.0)
        notes = st.text_input("Notes (optional)")
        submit_med = st.form_submit_button("Add Medicine")

    if submit_med:
        save_medicine_data({"Date": med_date, "Medicine": med_name, "Quantity": quantity, "Cost": cost, "Notes": notes})
        st.success("✅ Medicine added successfully!")

    st.header("📋 Medicine Records")
    med_df = load_medicines_data()
    if not med_df.empty:
        st.dataframe(med_df)
    else:
        st.info("No medicines recorded yet.")

# ---------------- Investments Tab ----------------
if menu == "💰 Investments":
    st.title("💰 Investments Tracker")

    with st.form("investment_form"):
        inv_date = st.date_input("Date", value=date.today())
        inv_type = st.selectbox("Investment Type", ["Salary", "SIP", "FD", "Stocks", "Chit Fund", "Other"])
        amount = st.number_input("Amount (₹)", min_value=0.0, step=100.0)
        frequency = st.selectbox("Frequency", ["One-time", "Monthly", "Quarterly"])
        notes = st.text_input("Notes (optional)")
        submit_inv = st.form_submit_button("Add Investment")

    if submit_inv:
        save_investment_data({"Date": inv_date, "Type": inv_type, "Amount": amount, "Frequency": frequency, "Notes": notes})
        st.success("✅ Investment added successfully!")

    st.header("📋 Investment Records")
    inv_df = load_investments_data()
    if not inv_df.empty:
        st.dataframe(inv_df)

        # 📊 Salary & Savings Summary
        inv_df["Date"] = pd.to_datetime(inv_df["Date"])
        inv_df["Month"] = inv_df["Date"].dt.strftime("%B %Y")
        selected_month = st.selectbox("📅 Select Month for Summary", sorted(inv_df["Month"].unique(), reverse=True))

        month_df = inv_df[inv_df["Month"] == selected_month]
        total_salary = month_df[month_df["Type"] == "Salary"]["Amount"].sum()
        total_investments = month_df[month_df["Type"] != "Salary"]["Amount"].sum()

        st.metric("💰 Total Salary", f"₹{total_salary:,.2f}")
        st.metric("📈 Total Investments", f"₹{total_investments:,.2f}")

        exp_df = load_expense_data()
        exp_df["Date"] = pd.to_datetime(exp_df["Date"])
        exp_df["Month"] = exp_df["Date"].dt.strftime("%B %Y")
        total_expenses = exp_df[exp_df["Month"] == selected_month]["Amount"].sum()

        savings = total_salary - (total_investments + total_expenses)
        st.metric("💵 Savings", f"₹{savings:,.2f}")
    else:
        st.info("No investments recorded yet.")
